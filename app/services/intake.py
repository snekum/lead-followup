"""Import walk-in leads from a daily CSV / Excel export and dedupe by phone.

The dealership keeps a fixed sheet; we ingest new rows, normalize phones, drop
duplicates (within the file and against existing leads), and report a summary.
"""
from __future__ import annotations

import csv
import datetime as dt
import io
from dataclasses import dataclass, field

from sqlalchemy import select
from sqlalchemy.orm import Session

from app.db.models import Event, Lead
from app.domain.enums import Language, LeadSource
from app.services.phone import PhoneError, normalize_phone

# Accepted header -> canonical field. Headers are lower-cased + underscored.
_HEADER_ALIASES: dict[str, str] = {
    "name": "name",
    "customer_name": "name",
    "customer": "name",
    "phone": "phone",
    "mobile": "phone",
    "phone_number": "phone",
    "mobile_number": "phone",
    "contact": "phone",
    "model": "interested_model",
    "interested_model": "interested_model",
    "car": "interested_model",
    "vehicle": "interested_model",
    "budget": "budget",
    "visit_date": "visit_date",
    "date_of_visit": "visit_date",
    "visit": "visit_date",
    "date": "visit_date",
    "language": "preferred_language",
    "lang": "preferred_language",
    "preferred_language": "preferred_language",
    "notes": "notes",
    "remarks": "notes",
}

_LANGUAGE_ALIASES: dict[str, Language] = {
    "en": Language.EN,
    "english": Language.EN,
    "hi": Language.HI,
    "hindi": Language.HI,
    "te": Language.TE,
    "telugu": Language.TE,
}

_DATE_FORMATS = ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%d.%m.%Y")


@dataclass
class RowError:
    row: int
    reason: str


@dataclass
class ImportResult:
    imported: int = 0
    duplicates: int = 0
    invalid: int = 0
    errors: list[RowError] = field(default_factory=list)

    def as_dict(self) -> dict[str, object]:
        return {
            "imported": self.imported,
            "duplicates": self.duplicates,
            "invalid": self.invalid,
            "errors": [{"row": e.row, "reason": e.reason} for e in self.errors],
        }


def _canonical_row(raw: dict[str, str]) -> dict[str, str]:
    out: dict[str, str] = {}
    for key, value in raw.items():
        if key is None:
            continue
        canonical = _HEADER_ALIASES.get(key.strip().lower().replace(" ", "_"))
        if canonical and value is not None and str(value).strip():
            out[canonical] = str(value).strip()
    return out


def _parse_date(value: str | None) -> dt.date | None:
    if not value:
        return None
    for fmt in _DATE_FORMATS:
        try:
            return dt.datetime.strptime(value, fmt).date()
        except ValueError:
            continue
    return None


def _parse_language(value: str | None) -> Language:
    if not value:
        return Language.EN
    return _LANGUAGE_ALIASES.get(value.strip().lower(), Language.EN)


def parse_rows(filename: str, content: bytes) -> list[dict[str, str]]:
    """Parse a CSV or XLSX file into a list of raw header->value dicts."""
    if filename.lower().endswith((".xlsx", ".xlsm")):
        return _parse_xlsx(content)
    return _parse_csv(content)


def _parse_csv(content: bytes) -> list[dict[str, str]]:
    text = content.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    return [dict(row) for row in reader]


def _parse_xlsx(content: bytes) -> list[dict[str, str]]:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    try:
        header = [str(c).strip() if c is not None else "" for c in next(rows)]
    except StopIteration:
        return []
    out: list[dict[str, str]] = []
    for values in rows:
        if values is None or all(v is None for v in values):
            continue
        out.append(
            {
                header[i]: ("" if v is None else str(v))
                for i, v in enumerate(values)
                if i < len(header)
            }
        )
    return out


def import_leads(
    session: Session,
    rows: list[dict[str, str]],
    *,
    source: LeadSource = LeadSource.CSV_IMPORT,
) -> ImportResult:
    """Create leads from parsed rows; dedupe by normalized phone."""
    result = ImportResult()

    # Normalize first so we can batch-check existing phones.
    prepared: list[tuple[int, dict[str, str], str]] = []
    for idx, raw in enumerate(rows, start=2):  # row 1 is the header
        row = _canonical_row(raw)
        if not row.get("name") or not row.get("phone"):
            result.invalid += 1
            result.errors.append(RowError(idx, "missing name or phone"))
            continue
        try:
            phone = normalize_phone(row["phone"])
        except PhoneError as exc:
            result.invalid += 1
            result.errors.append(RowError(idx, str(exc)))
            continue
        prepared.append((idx, row, phone))

    if not prepared:
        return result

    candidate_phones = {phone for _, _, phone in prepared}
    existing: set[str] = set(
        session.scalars(select(Lead.phone).where(Lead.phone.in_(candidate_phones)))
    )

    seen_in_file: set[str] = set()
    for _idx, row, phone in prepared:
        if phone in existing or phone in seen_in_file:
            result.duplicates += 1
            continue
        seen_in_file.add(phone)
        lead = Lead(
            name=row["name"],
            phone=phone,
            source=source,
            interested_model=row.get("interested_model"),
            budget=row.get("budget"),
            visit_date=_parse_date(row.get("visit_date")),
            preferred_language=_parse_language(row.get("preferred_language")),
            notes=row.get("notes"),
        )
        session.add(lead)
        session.flush()
        session.add(
            Event(lead_id=lead.id, type="lead_imported", data={"source": source.value})
        )
        result.imported += 1

    session.commit()
    return result


def import_file(session: Session, filename: str, content: bytes) -> ImportResult:
    rows = parse_rows(filename, content)
    return import_leads(session, rows)
